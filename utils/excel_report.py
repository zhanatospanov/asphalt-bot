"""Генерация Excel-отчёта по рейсам."""
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "№ накладной", "Дата", "Время", "Гос. номер авто",
    "Покупатель", "Объект", "Марка асфальта",
    "Тара, т", "Брутто, т", "Нетто, т"
]

COL_WIDTHS = [14, 12, 8, 16, 28, 32, 28, 10, 10, 10]


def _border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill():
    return PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")


def generate_excel(trips: list, date_from: str = None, date_to: str = None,
                   company_name: str = "") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Журнал отпуска"

    # Заголовок отчёта
    title = f"Журнал отпуска продукции"
    if company_name:
        title += f"  —  {company_name}"
    ws.merge_cells("A1:I1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center")

    period = ""
    if date_from and date_to:
        period = f"Период: {_fmt(date_from)} — {_fmt(date_to)}"
    elif date_from:
        period = f"С {_fmt(date_from)}"
    elif date_to:
        period = f"По {_fmt(date_to)}"
    else:
        period = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"

    ws.merge_cells("A2:I2")
    ws["A2"] = period
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Заголовки таблицы
    row = 4
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[row].height = 30

    # Данные
    total_net = 0.0
    total_gross = 0.0
    alt_fill = PatternFill(start_color="EBF2FA", end_color="EBF2FA", fill_type="solid")

    for i, trip in enumerate(trips, start=1):
        r = row + i
        net = trip.get("net_kg", 0)
        gross = trip.get("gross_kg", 0)
        tare = trip.get("tare_kg", 0)
        total_net += net
        total_gross += gross

        values = [
            trip.get("doc_number", ""),
            _fmt(trip.get("trip_date", "")),
            trip.get("trip_time", ""),
            trip.get("vehicle_number", ""),
            trip.get("buyer_name", ""),
            trip.get("object_name", ""),
            trip.get("asphalt_grade", ""),
            round(tare / 1000, 3),
            round(gross / 1000, 3),
            round(net / 1000, 3),
        ]

        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = _border()
            cell.font = Font(size=9)
            align = "center" if col in (1, 2, 3, 4, 8, 9, 10) else "left"
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if fill:
                cell.fill = fill

            # Числа с точностью
            if col in (8, 9, 10):
                cell.number_format = "0.000"

    # Итоговая строка
    total_row = row + len(trips) + 1
    ws.merge_cells(f"A{total_row}:G{total_row}")
    total_cell = ws[f"A{total_row}"]
    total_cell.value = "ИТОГО:"
    total_cell.font = Font(bold=True, size=9)
    total_cell.alignment = Alignment(horizontal="right")
    total_cell.border = _border()

    tare_total = total_gross - total_net
    for col, val in [(8, round(tare_total / 1000, 3)),
                     (9, round(total_gross / 1000, 3)),
                     (10, round(total_net / 1000, 3))]:
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True, size=9)
        cell.number_format = "0.000"
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Подпись
    sign_row = total_row + 2
    ws[f"A{sign_row}"] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws[f"A{sign_row}"].font = Font(italic=True, size=8, color="888888")

    ws.freeze_panes = f"A{row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fmt(date_str):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return date_str
